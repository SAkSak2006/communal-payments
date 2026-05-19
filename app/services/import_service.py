import io
from decimal import Decimal, InvalidOperation
from typing import List, Dict

from openpyxl import Workbook, load_workbook
from sqlalchemy import select
from sqlalchemy.ext.asyncio import AsyncSession

from app.models import Resident, Address, Meter, UtilityService, MeterReading, ReadingSource


async def import_residents_from_excel(
    session: AsyncSession, file_content: bytes
) -> Dict:
    """Import residents from Excel file.

    Expected columns:
    A: ФИО
    B: Телефон
    C: Лицевой счёт
    D: Город
    E: Улица
    F: Дом
    G: Квартира
    H: Площадь (м²)
    """
    wb = load_workbook(io.BytesIO(file_content), read_only=True)
    ws = wb.active

    created = 0
    skipped = 0
    errors = []

    for row_idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
        if not row or not row[0]:
            continue

        try:
            full_name = str(row[0]).strip()
            phone = str(row[1]).strip() if row[1] else ""
            account = str(row[2]).strip() if row[2] else ""
            city = str(row[3]).strip() if row[3] else ""
            street = str(row[4]).strip() if row[4] else ""
            building = str(row[5]).strip() if row[5] else ""
            apartment = str(row[6]).strip() if row[6] else ""
            area = Decimal(str(row[7]).strip()) if row[7] else Decimal("0")

            if not all([full_name, phone, account, city, street, building, apartment]):
                errors.append(f"Строка {row_idx}: не все обязательные поля заполнены")
                skipped += 1
                continue

            # Check duplicate account
            existing = await session.execute(
                select(Resident).where(Resident.personal_account == account)
            )
            if existing.scalar_one_or_none():
                skipped += 1
                continue

            # Check duplicate phone
            existing_phone = await session.execute(
                select(Resident).where(Resident.phone == phone)
            )
            if existing_phone.scalar_one_or_none():
                errors.append(f"Строка {row_idx}: телефон {phone} уже существует")
                skipped += 1
                continue

            resident = Resident(
                full_name=full_name,
                phone=phone,
                personal_account=account,
                is_verified=True,
            )
            session.add(resident)
            await session.flush()

            address = Address(
                city=city, street=street, building=building, apartment=apartment,
                area_sqm=area, resident_id=resident.id,
            )
            session.add(address)
            await session.flush()

            # Auto-create meters for services that have meters
            result = await session.execute(
                select(UtilityService).where(UtilityService.has_meter.is_(True))
            )
            meter_services = result.scalars().all()
            for svc in meter_services:
                meter = Meter(
                    address_id=address.id,
                    service_id=svc.id,
                    serial_number=f"{svc.code[:2].upper()}-{account}-{svc.id}",
                )
                session.add(meter)

            created += 1

        except Exception as e:
            errors.append(f"Строка {row_idx}: {str(e)}")
            skipped += 1

    await session.commit()
    wb.close()

    return {"created": created, "skipped": skipped, "errors": errors}


async def import_readings_from_excel(
    session: AsyncSession, file_content: bytes
) -> Dict:
    """Import meter readings from Excel file.

    Expected columns:
    A: Лицевой счёт
    B: Код услуги (cold_water, hot_water, electricity, gas)
    C: Показание
    D: Год
    E: Месяц
    """
    wb = load_workbook(io.BytesIO(file_content), read_only=True)
    ws = wb.active

    created = 0
    skipped = 0
    errors = []

    for row_idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
        if not row or not row[0]:
            continue

        try:
            account = str(row[0]).strip()
            service_code = str(row[1]).strip() if row[1] else ""
            value = Decimal(str(row[2]).strip()) if row[2] else None
            year = int(row[3]) if row[3] else None
            month = int(row[4]) if row[4] else None

            if not all([account, service_code, value is not None, year, month]):
                errors.append(f"Строка {row_idx}: не все поля заполнены")
                skipped += 1
                continue

            # Find resident
            res = await session.execute(
                select(Resident).where(Resident.personal_account == account)
            )
            resident = res.scalar_one_or_none()
            if not resident:
                errors.append(f"Строка {row_idx}: счёт {account} не найден")
                skipped += 1
                continue

            # Find service
            res = await session.execute(
                select(UtilityService).where(UtilityService.code == service_code)
            )
            service = res.scalar_one_or_none()
            if not service:
                errors.append(f"Строка {row_idx}: услуга {service_code} не найдена")
                skipped += 1
                continue

            # Find meter
            res = await session.execute(
                select(Meter)
                .join(Address, Address.id == Meter.address_id)
                .where(
                    Address.resident_id == resident.id,
                    Meter.service_id == service.id,
                    Meter.is_active.is_(True),
                )
            )
            meter = res.scalar_one_or_none()
            if not meter:
                errors.append(f"Строка {row_idx}: счётчик не найден для {account}/{service_code}")
                skipped += 1
                continue

            # Check duplicate reading
            res = await session.execute(
                select(MeterReading).where(
                    MeterReading.meter_id == meter.id,
                    MeterReading.period_year == year,
                    MeterReading.period_month == month,
                )
            )
            if res.scalar_one_or_none():
                skipped += 1
                continue

            # Get previous reading
            res = await session.execute(
                select(MeterReading)
                .where(MeterReading.meter_id == meter.id)
                .order_by(MeterReading.reading_date.desc())
                .limit(1)
            )
            last = res.scalar_one_or_none()
            prev_value = last.value if last else Decimal("0")
            consumption = max(Decimal("0"), value - prev_value)

            reading = MeterReading(
                meter_id=meter.id,
                value=value,
                previous_value=prev_value,
                consumption=consumption,
                period_year=year,
                period_month=month,
                source=ReadingSource.ADMIN,
                is_validated=True,
            )
            session.add(reading)
            created += 1

        except Exception as e:
            errors.append(f"Строка {row_idx}: {str(e)}")
            skipped += 1

    await session.commit()
    wb.close()

    return {"created": created, "skipped": skipped, "errors": errors}


def generate_residents_template() -> io.BytesIO:
    """Generate Excel template for residents import."""
    wb = Workbook()
    ws = wb.active
    ws.title = "Жители"

    headers = ["ФИО", "Телефон", "Лицевой счёт", "Город", "Улица", "Дом", "Квартира", "Площадь (м²)"]
    for col, h in enumerate(headers, 1):
        ws.cell(row=1, column=col, value=h)

    # Example row
    example = ["Иванов Иван Иванович", "+79001234567", "1001-0001", "Москва", "Ленина", "10", "1", 54.5]
    for col, v in enumerate(example, 1):
        ws.cell(row=2, column=col, value=v)

    for col in range(1, 9):
        ws.column_dimensions[chr(64 + col)].width = 20

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    return output


def generate_readings_template() -> io.BytesIO:
    """Generate Excel template for readings import."""
    wb = Workbook()
    ws = wb.active
    ws.title = "Показания"

    headers = ["Лицевой счёт", "Код услуги", "Показание", "Год", "Месяц"]
    for col, h in enumerate(headers, 1):
        ws.cell(row=1, column=col, value=h)

    examples = [
        ["1001-0001", "cold_water", 150.5, 2025, 3],
        ["1001-0001", "hot_water", 85.2, 2025, 3],
        ["1001-0001", "electricity", 1250, 2025, 3],
    ]
    for row_idx, ex in enumerate(examples, 2):
        for col, v in enumerate(ex, 1):
            ws.cell(row=row_idx, column=col, value=v)

    # Add note about service codes
    ws2 = wb.create_sheet("Коды услуг")
    ws2.cell(row=1, column=1, value="Код")
    ws2.cell(row=1, column=2, value="Услуга")
    codes = [
        ("cold_water", "Холодное водоснабжение"),
        ("hot_water", "Горячее водоснабжение"),
        ("electricity", "Электроснабжение"),
        ("gas", "Газоснабжение"),
    ]
    for i, (code, name) in enumerate(codes, 2):
        ws2.cell(row=i, column=1, value=code)
        ws2.cell(row=i, column=2, value=name)

    for col in range(1, 6):
        ws.column_dimensions[chr(64 + col)].width = 20

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    return output
