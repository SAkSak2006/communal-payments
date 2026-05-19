import io
from typing import List, Dict

from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill


def export_charges_to_excel(
    charges: List[Dict], year: int, month: int
) -> io.BytesIO:
    """Export charges to Excel file."""
    wb = Workbook()
    ws = wb.active
    ws.title = f"Начисления {month:02d}.{year}"

    # Styles
    header_font = Font(bold=True, size=12)
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_font_white = Font(bold=True, size=11, color="FFFFFF")
    thin_border = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin"),
    )

    # Title
    ws.merge_cells("A1:F1")
    ws["A1"] = f"Начисления за {month:02d}.{year}"
    ws["A1"].font = header_font
    ws["A1"].alignment = Alignment(horizontal="center")

    # Headers
    headers = ["Житель", "Лиц. счёт", "Адрес", "Услуга", "Потребление", "Сумма (руб.)"]
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=3, column=col, value=header)
        cell.font = header_font_white
        cell.fill = header_fill
        cell.border = thin_border
        cell.alignment = Alignment(horizontal="center")

    # Data
    total = 0
    for i, charge in enumerate(charges, 4):
        ws.cell(row=i, column=1, value=charge["resident_name"]).border = thin_border
        ws.cell(row=i, column=2, value=charge["personal_account"]).border = thin_border
        ws.cell(row=i, column=3, value=charge["address"]).border = thin_border
        ws.cell(row=i, column=4, value=charge["service_name"]).border = thin_border
        cons_cell = ws.cell(row=i, column=5, value=float(charge["consumption"]))
        cons_cell.border = thin_border
        cons_cell.number_format = "0.000"
        amt_cell = ws.cell(row=i, column=6, value=float(charge["amount"]))
        amt_cell.border = thin_border
        amt_cell.number_format = "#,##0.00"
        total += float(charge["amount"])

    # Total row
    total_row = len(charges) + 4
    ws.cell(row=total_row, column=5, value="ИТОГО:").font = Font(bold=True)
    total_cell = ws.cell(row=total_row, column=6, value=total)
    total_cell.font = Font(bold=True)
    total_cell.number_format = "#,##0.00"

    # Column widths
    ws.column_dimensions["A"].width = 25
    ws.column_dimensions["B"].width = 15
    ws.column_dimensions["C"].width = 30
    ws.column_dimensions["D"].width = 25
    ws.column_dimensions["E"].width = 15
    ws.column_dimensions["F"].width = 15

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    return output


def export_debtors_to_excel(debtors: List[Dict]) -> io.BytesIO:
    """Export debtors report to Excel."""
    wb = Workbook()
    ws = wb.active
    ws.title = "Должники"

    header_font = Font(bold=True, size=12)
    header_fill = PatternFill(start_color="C00000", end_color="C00000", fill_type="solid")
    header_font_white = Font(bold=True, size=11, color="FFFFFF")
    thin_border = Border(
        left=Side(style="thin"), right=Side(style="thin"),
        top=Side(style="thin"), bottom=Side(style="thin"),
    )

    ws.merge_cells("A1:E1")
    ws["A1"] = "Отчёт по задолженностям"
    ws["A1"].font = header_font
    ws["A1"].alignment = Alignment(horizontal="center")

    headers = ["Житель", "Лиц. счёт", "Начислено", "Оплачено", "Задолженность"]
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=3, column=col, value=header)
        cell.font = header_font_white
        cell.fill = header_fill
        cell.border = thin_border

    for i, d in enumerate(debtors, 4):
        ws.cell(row=i, column=1, value=d["full_name"]).border = thin_border
        ws.cell(row=i, column=2, value=d["personal_account"]).border = thin_border
        c = ws.cell(row=i, column=3, value=float(d["charged"]))
        c.border = thin_border
        c.number_format = "#,##0.00"
        p = ws.cell(row=i, column=4, value=float(d["paid"]))
        p.border = thin_border
        p.number_format = "#,##0.00"
        debt_cell = ws.cell(row=i, column=5, value=float(d["debt"]))
        debt_cell.border = thin_border
        debt_cell.number_format = "#,##0.00"
        debt_cell.font = Font(bold=True, color="C00000")

    ws.column_dimensions["A"].width = 25
    ws.column_dimensions["B"].width = 15
    ws.column_dimensions["C"].width = 15
    ws.column_dimensions["D"].width = 15
    ws.column_dimensions["E"].width = 18

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    return output
