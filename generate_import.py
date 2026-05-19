#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""Generate import files: 1000 residents + 1000 readings."""

import random
from openpyxl import Workbook

random.seed(42)

# === DATA ===
lastnames = ['Иванов','Петров','Сидоров','Козлов','Новиков','Морозов','Волков','Соколов','Попов','Лебедев',
             'Кузнецов','Смирнов','Васильев','Федоров','Николаев','Алексеев','Павлов','Семёнов','Григорьев',
             'Степанов','Белов','Тарасов','Комаров','Орлов','Киселёв','Макаров','Андреев','Ковалёв','Захаров',
             'Борисов','Герасимов','Пономарёв','Романов','Осипов','Егоров','Баранов','Беляев','Рябов',
             'Калинин','Сергеев','Антонов','Тимофеев','Никитин','Крылов','Максимов','Мельников','Дмитриев',
             'Назаров','Филиппов','Веселов']

male_names = ['Иван','Пётр','Алексей','Дмитрий','Сергей','Андрей','Михаил','Николай','Александр',
              'Владимир','Артём','Максим','Евгений','Олег','Виктор','Константин','Юрий','Роман',
              'Валерий','Игорь','Денис','Антон','Борис','Василий','Павел','Кирилл','Фёдор','Тимур','Руслан']

female_names = ['Анна','Мария','Елена','Ольга','Наталья','Ирина','Татьяна','Светлана','Екатерина',
                'Дарья','Юлия','Алина','Виктория','Кристина','Марина','Вера','Галина','Людмила',
                'Надежда','Полина','Валентина','Лариса','Софья','Диана','Оксана','Нина','Зоя','Антонина']

male_patro = ['Иванович','Петрович','Алексеевич','Дмитриевич','Сергеевич','Андреевич','Михайлович',
              'Николаевич','Александрович','Владимирович','Евгеньевич','Олегович','Викторович',
              'Константинович','Юрьевич','Романович','Валерьевич','Игоревич','Денисович','Павлович']

female_patro = ['Ивановна','Петровна','Алексеевна','Дмитриевна','Сергеевна','Андреевна','Михайловна',
                'Николаевна','Александровна','Владимировна','Евгеньевна','Олеговна','Викторовна',
                'Константиновна','Юрьевна','Романовна','Валерьевна','Игоревна','Денисовна','Павловна']

cities = ['Москва','Санкт-Петербург','Казань','Новосибирск','Екатеринбург',
          'Краснодар','Нижний Новгород','Самара','Ростов-на-Дону','Воронеж']

streets = ['Ленина','Пушкина','Гагарина','Мира','Советская','Кирова','Горького','Чехова','Лесная',
           'Садовая','Молодёжная','Центральная','Школьная','Полевая','Набережная','Космонавтов',
           'Победы','Труда','Парковая','Берёзовая','Солнечная','Речная','Озёрная','Северная','Южная']

prefixes = ['900','901','902','903','904','905','906','910','911','912','913','914','915','916',
            '917','918','919','920','921','922','923','925','926','927','928','929','950','951',
            '952','960','961','962','963','964','965','977','980','985','987','988','999']


def make_female_lastname(ln):
    if ln.endswith('ов'): return ln[:-2] + 'ова'
    if ln.endswith('ёв'): return ln[:-2] + 'ёва'
    if ln.endswith('ев'): return ln[:-2] + 'ева'
    if ln.endswith('ин'): return ln[:-2] + 'ина'
    return ln + 'а'


# ============ 1. RESIDENTS (1000) ============
wb = Workbook()
ws = wb.active
ws.title = "Жители"

headers = ["ФИО", "Телефон", "Лицевой счёт", "Город", "Улица", "Дом", "Квартира", "Площадь (м²)"]
for c, h in enumerate(headers, 1):
    ws.cell(1, c, h)

used_phones = set()
for i in range(1, 1001):
    is_male = random.random() < 0.5
    ln = random.choice(lastnames)
    if not is_male:
        ln = make_female_lastname(ln)
    fn = random.choice(male_names if is_male else female_names)
    pt = random.choice(male_patro if is_male else female_patro)
    fio = f"{ln} {fn} {pt}"

    while True:
        phone = f"+7{random.choice(prefixes)}{random.randint(1000000, 9999999)}"
        if phone not in used_phones:
            used_phones.add(phone)
            break

    acc = f"LS-{1000000 + i}"
    city = random.choice(cities)
    street = f"ул. {random.choice(streets)}"
    bld = str(random.randint(1, 150))
    apt = str(random.randint(1, 300))
    area = round(random.uniform(25, 120), 1)

    # Guarantee unique address: use account number as apartment
    apt = str(i)
    row = [fio, phone, acc, city, street, bld, apt, area]
    for c, v in enumerate(row, 1):
        ws.cell(i + 1, c, v)

for c in range(1, 9):
    ws.column_dimensions[chr(64 + c)].width = 22

wb.save("import_zhiteli_1000.xlsx")


# ============ 2. READINGS (1000) ============
wb2 = Workbook()
ws2 = wb2.active
ws2.title = "Показания"

headers2 = ["Лицевой счёт", "Код услуги", "Показание", "Год", "Месяц"]
for c, h in enumerate(headers2, 1):
    ws2.cell(1, c, h)

svcs = ["cold_water", "hot_water", "electricity", "gas"]
ranges_map = {
    "cold_water": (50, 300),
    "hot_water": (30, 200),
    "electricity": (500, 5000),
    "gas": (20, 150),
}

accounts = [f"LS-{1000000 + i}" for i in range(1, 1001)]
random.shuffle(accounts)

for i in range(1000):
    acc = accounts[i]
    svc = random.choice(svcs)
    lo, hi = ranges_map[svc]
    val = round(random.uniform(lo, hi), 1)
    yr = random.choice([2024, 2025, 2025, 2025])
    mn = random.randint(1, 12) if yr == 2024 else random.randint(1, 3)

    ws2.cell(i + 2, 1, acc)
    ws2.cell(i + 2, 2, svc)
    ws2.cell(i + 2, 3, val)
    ws2.cell(i + 2, 4, yr)
    ws2.cell(i + 2, 5, mn)

# Codes reference sheet
ws3 = wb2.create_sheet("Коды услуг")
ws3.cell(1, 1, "Код")
ws3.cell(1, 2, "Услуга")
codes = [
    ("cold_water", "Холодное водоснабжение"),
    ("hot_water", "Горячее водоснабжение"),
    ("electricity", "Электроснабжение"),
    ("gas", "Газоснабжение"),
]
for j, (c, n) in enumerate(codes, 2):
    ws3.cell(j, 1, c)
    ws3.cell(j, 2, n)

for c in range(1, 6):
    ws2.column_dimensions[chr(64 + c)].width = 20

wb2.save("import_pokazaniya_1000.xlsx")

print("OK")
